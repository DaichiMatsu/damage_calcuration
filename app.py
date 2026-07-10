import tkinter 
from tkinter import ttk
from tkinter import filedialog
import openpyxl
from pathlib import Path
import math

class Application(tkinter.Frame):
    def __init__(self, root=None):
        super().__init__(root,
                         width=820, height=1120,
                         borderwidth=1, relief="groove")
        self.root = root
        self.pack() # 位置を設定して配置
        self.pack_propagate(0) # サイズ調整
        self.create_widgets()


    def create_widgets(self):
        # 閉じるボタン
        quit_btn = tkinter.Button(self)
        quit_btn["text"] = "閉じる"
        quit_btn["command"] = self.root.destroy
        quit_btn.pack(side="bottom") # 位置指定　bottom=下

        self.cell_positions = []
        self.cell_texts = []

        self.message = tkinter.Message(self)
        self.message.place(x="180", y=310)

        self.loading_message = tkinter.Message(self)
        self.loading_message.place(x="180", y=310)

    
        # メッセージ出力(A/C)
        text_label = tkinter.Message(self)
        text_label["text"] = "A/C"
        text_label.place(x=10, y=str(22))
        # テキストボックス(A/C)
        self.text_box_A_C = tkinter.Entry(self)
        self.text_box_A_C["width"] = 10
        self.text_box_A_C.place(x="50", y=str(22))
        self.cell_texts.append(self.text_box_A_C)   

        # メッセージ出力(努力値A/C)
        text_label = tkinter.Message(self)
        text_label["text"] = "A/C努力値"
        text_label.place(x=10, y=str(2*22))
        # テキストボックス(努力値A/C)
        self.text_box_efA_C = tkinter.Entry(self)
        self.text_box_efA_C["width"] = 10
        self.text_box_efA_C.place(x="50", y=str(2*22))
        self.cell_texts.append(self.text_box_efA_C) 

        # メッセージ出力(わざ)
        text_label = tkinter.Message(self)
        text_label["text"] = "わざ"
        text_label.place(x=10, y=str(3*22))
        # テキストボックス(わざ)
        self.text_box_mo = tkinter.Entry(self)
        self.text_box_mo["width"] = 10
        self.text_box_mo.place(x="50", y=str(3*22))
        self.cell_texts.append(self.text_box_mo)   

        # メッセージ出力(タイプ一致)
        text_label = tkinter.Message(self)
        text_label["text"] = "タイプ一致"
        text_label.place(x=10, y=str(4*22))
        # チェックボックス(タイプ一致)
        self.checked_type_same = tkinter.BooleanVar()
        self.check_box_type_same = tkinter.Checkbutton(
            self,
            variable=self.checked_type_same
        )
        self.check_box_type_same.place(x=100, y=100)
        self.check_box_type_same.place(x="50", y=str(4*22))

        # メッセージ出力(性格AC)
        text_label = tkinter.Message(self)
        text_label["text"] = "性格"
        text_label.place(x=10, y=str(5*22))
        # テキストボックス(性格AC)
        self.text_box_nature_AC = ttk.Combobox(
        self,
        values=["+", "0", "-"],  # 選択肢
        state="readonly",           # 入力不可（選択のみ）
        width=10
        )
        self.text_box_nature_AC.current(1)  # 0を選択状態にする
        self.text_box_nature_AC.place(x="50", y=str(5*22))

        # メッセージ出力(ランクAC)
        text_label = tkinter.Message(self)
        text_label["text"] = "ランクAC"
        text_label.place(x=10, y=str(6*22))
        # テキストボックス(ランクAC)
        self.text_box_rank_AC = ttk.Combobox(
        self,
        values=["-6", "-5", "-4", "-3", "-2", "-1", "0", "1", "2", "3", "4", "5", "6"],  # 選択肢
        state="readonly",           # 入力不可（選択のみ）
        width=10
        )
        self.text_box_rank_AC.current(6)  # 0を選択状態にする
        self.text_box_rank_AC.place(x="50", y=str(6*22))


        # メッセージ出力(B/D)
        text_label = tkinter.Message(self)
        text_label["text"] = "B/D"
        text_label.place(x=200, y=str(22))
        # テキストボックス(B/D)
        self.text_box_B_D = tkinter.Entry(self)
        self.text_box_B_D["width"] = 10
        self.text_box_B_D.place(x="240", y=str(22))
        self.cell_texts.append(self.text_box_B_D)   

        # メッセージ出力(HP)
        text_label = tkinter.Message(self)
        text_label["text"] = "HP"
        text_label.place(x=200, y=str(2*22))
        # テキストボックス(HP)
        self.text_box_HP = tkinter.Entry(self)
        self.text_box_HP["width"] = 10
        self.text_box_HP.place(x="240", y=str(2*22))
        self.cell_texts.append(self.text_box_HP) 

        # メッセージ出力(努力値B/D)
        text_label = tkinter.Message(self)
        text_label["text"] = "B/D努力値"
        text_label.place(x=200, y=str(3*22))
        # テキストボックス(努力値B/D)
        self.text_box_efB_D = tkinter.Entry(self)
        self.text_box_efB_D["width"] = 10
        self.text_box_efB_D.place(x="240", y=str(3*22))
        self.cell_texts.append(self.text_box_efB_D) 

        # メッセージ出力(努力値HP)
        text_label = tkinter.Message(self)
        text_label["text"] = "HP努力値"
        text_label.place(x=200, y=str(4*22))
        # テキストボックス(努力値HP)
        self.text_box_efHP = tkinter.Entry(self)
        self.text_box_efHP["width"] = 10
        self.text_box_efHP.place(x="240", y=str(4*22))
        self.cell_texts.append(self.text_box_efHP) 

        # メッセージ出力(タイプ相性)
        text_label = tkinter.Message(self)
        text_label["text"] = "タイプ相性"
        text_label.place(x=200, y=str(5*22))
        # テキストボックス(タイプ相性)
        self.text_box_type_match = ttk.Combobox(
        self,
        values=["等倍", "ちょうばつぐん", "ばつぐん", "いまひとつ", "かなりいまひとつ", "こうかなし"],  # 選択肢
        state="readonly",           # 入力不可（選択のみ）
        width=10
        )
        self.text_box_type_match.current(0)  # 最初の項目を選択状態にする
        self.text_box_type_match.place(x="240", y=str(5*22))

        # メッセージ出力(性格BD)
        text_label = tkinter.Message(self)
        text_label["text"] = "性格"
        text_label.place(x=200, y=str(6*22))
        # テキストボックス(性格BD)
        self.text_box_nature_BD = ttk.Combobox(
        self,
        values=["+", "0", "-"],  # 選択肢
        state="readonly",           # 入力不可（選択のみ）
        width=10
        )
        self.text_box_nature_BD.current(1)  # 0を選択状態にする
        self.text_box_nature_BD.place(x="240", y=str(6*22))

        # メッセージ出力(ランクBD)
        text_label = tkinter.Message(self)
        text_label["text"] = "ランクBD"
        text_label.place(x=200, y=str(7*22))
        # テキストボックス(ランクBD)
        self.text_box_rank_BD = ttk.Combobox(
        self,
        values=["-6", "-5", "-4", "-3", "-2", "-1", "0", "1", "2", "3", "4", "5", "6"],  # 選択肢
        state="readonly",           # 入力不可（選択のみ）
        width=10
        )
        self.text_box_rank_BD.current(6)  # ±0を選択状態にする
        self.text_box_rank_BD.place(x="240", y=str(7*22))


        # 実行ボタン
        submit_btn = tkinter.Button(self)
        submit_btn["text"] = "保存"
        submit_btn["command"] = self.save_data
        submit_btn.place(x=150, y=230)

        # 計算ボタン
        calc_btn = tkinter.Button(self)
        calc_btn["text"] = "計算"
        calc_btn["command"] = self.calculation
        calc_btn.place(x=150, y=600)

        # リセットボタン
        reset_btn = tkinter.Button(self)
        reset_btn["text"] = "リセット"
        reset_btn["command"] = self.reset
        reset_btn.place(x=150, y=400)
        

        # 読み込みボタン
        submit_btn = tkinter.Button(self)
        submit_btn["text"] = "読み込み"
        submit_btn["command"] = self.loading_data
        submit_btn.place(x=150, y=270)

    # 実行ボタンを押すとテキストを入力するシステムをメソッドとして作成
    def save_data(self):
        wb = openpyxl.load_workbook("app_data.xlsx") # 選択したファイルを読み込み
        ws = wb.worksheets[0] # ワークシートの一枚目(0番目)をwsに保存
        text_A_C = self.text_box_A_C.get()
        ws["A1"].value = text_A_C
        text_efA_C = self.text_box_efA_C.get()
        ws["A2"].value = text_efA_C
        text_mo = self.text_box_mo.get()
        ws["A3"].value = text_mo
        check_box_type_same = self.checked_type_same.get()
        ws["A4"].value = check_box_type_same
        text_box_nature_AC = self.text_box_nature_AC.get()
        ws["A5"].value = text_box_nature_AC
        text_box_rank_AC = self.text_box_rank_AC.get()
        ws["A6"].value = text_box_rank_AC


        text_B_D = self.text_box_B_D.get()
        ws["B1"].value = text_B_D
        text_HP = self.text_box_HP.get()
        ws["B2"].value = text_HP
        text_efB_D = self.text_box_efB_D.get()
        ws["B3"].value = text_efB_D
        text_efHP = self.text_box_efHP.get()
        ws["B4"].value = text_efHP
        text_box_type_match = self.text_box_type_match.get()
        ws["B5"].value = text_box_type_match
        text_box_nature_BD = self.text_box_nature_BD.get()
        ws["B6"].value = text_box_nature_BD
        text_box_rank_BD = self.text_box_rank_BD.get()
        ws["B7"].value = text_box_rank_BD
        
        
        wb.save("app_data.xlsx") # book保存
        self.message["text"] = "保存完了"


        # 五捨五超入
    def gosyagotyonyu(self, x):
        return math.ceil(x) if x - math.floor(x) > 0.5 else math.floor(x)
    
    def calculation(self):
        wb = openpyxl.load_workbook("app_data.xlsx")
        ws = wb.worksheets[0]

        # 文字列をint型に直す
        A_C = int(ws["A1"].value)
        efA_C = int(ws["A2"].value)
        mo = int(ws["A3"].value)
        B_D = int(ws["B1"].value)
        HP = int(ws["B2"].value)
        efB_D = int(ws["B3"].value)
        efHP = int(ws["B4"].value)

        # 実数値(HP)
        HP_act = math.floor(((((HP-75)*2) + 31 + (efHP*2)) * 50 / 100) + 60)

        # 実数値(こうげき、ぼうぎょ)
        A_C_act = math.floor(((((A_C-20)*2) + 31 + (efA_C*2)) * 50 / 100) + 5)
        B_D_act = math.floor(((((B_D-20)*2) + 31 + (efB_D*2)) * 50 / 100) + 5)

        # 性格補正
        if ws["A5"].value == "+":
            A_C_act_nat = math.floor(A_C_act * 1.1) 
        elif ws["A5"].value == "-":
            A_C_act_nat = math.floor(A_C_act * 0.9)
        else:
            A_C_act_nat = A_C_act

        if ws["B6"].value == "+":
            B_D_act_nat = math.floor(B_D_act * 1.1) 
        elif ws["B6"].value == "-":
            B_D_act_nat = math.floor(B_D_act * 0.9)
        else:
            B_D_act_nat = B_D_act   

        # ランク補正
        if int(ws["A6"].value) >= 0:
            A_C_rank = math.floor(A_C_act_nat * (2+int(ws["A6"].value))/2)
        else:
            A_C_rank = math.floor(A_C_act_nat * 2/(2+int(ws["A6"].value)*-1))

        if int(ws["B7"].value) >= 0:
            B_D_rank = math.floor(B_D_act_nat * (2+int(ws["B7"].value))/2)
        else:
            B_D_rank = math.floor(B_D_act_nat * 2/(2+int(ws["B7"].value)*-1))

        # 素のダメージ
        base_damage = math.floor(22 * mo * A_C_rank / B_D_rank)

        # 実ダメージ
        damage = math.floor((base_damage / 50) + 2)
        # 急所
        damage_cr = self.gosyagotyonyu(damage * 1.5)

        # 乱数計算(最小値)
        rand_damage = math.floor(damage * 85/100)
        # 急所
        rand_damage_cr = math.floor(damage_cr * 85/100)

        # タイプ一致
        if int(ws["A4"].value) == True:
            # 最大
            # 五捨五超入
            sametype_damage_max = self.gosyagotyonyu(damage * 1.5)
            # 最小
            # 五捨五超入
            sametype_damage_min = self.gosyagotyonyu(rand_damage * 1.5)

            # 急所
            # 最大
            # 五捨五超入
            sametype_damage_max_cr = self.gosyagotyonyu(damage_cr * 1.5)
            # 最小
            # 五捨五超入
            sametype_damage_min_cr = self.gosyagotyonyu(rand_damage_cr * 1.5)
        else:
            # 最大
            sametype_damage_max = damage
            # 最小
            sametype_damage_min = rand_damage

            # 急所
            # 最大
            sametype_damage_max_cr = damage_cr
            # 最小
            sametype_damage_min_cr = rand_damage_cr
        
        # タイプ相性
        if ws["B5"].value == "ちょうばつぐん":
            # 最大
            # 切り捨て
            match_damage_max = math.floor(sametype_damage_max * 4)
            # 最小
            # 切り捨て
            match_damage_min = math.floor(sametype_damage_min * 4)

            # 急所
            # 最大
            # 切り捨て
            match_damage_max_cr = math.floor(sametype_damage_max_cr * 4)
            # 最小
            # 切り捨て
            match_damage_min_cr = math.floor(sametype_damage_min_cr * 4)
        elif ws["B5"].value == "ばつぐん":
            # 最大
            # 切り捨て
            match_damage_max = math.floor(sametype_damage_max * 2)
            # 最小
            # 切り捨て
            match_damage_min = math.floor(sametype_damage_min * 2)

            # 急所
            # 最大
            # 切り捨て
            match_damage_max_cr = math.floor(sametype_damage_max_cr * 2)
            # 最小
            # 切り捨て
            match_damage_min_cr = math.floor(sametype_damage_min_cr * 2)
        elif ws["B5"].value == "いまひとつ":
            # 最大
            # 切り捨て
            match_damage_max = math.floor(sametype_damage_max * 0.5)
            # 最小
            # 切り捨て
            match_damage_min = math.floor(sametype_damage_min * 0.5)

            # 急所
            # 最大
            # 切り捨て
            match_damage_max_cr = math.floor(sametype_damage_max_cr * 0.5)
            # 最小
            # 切り捨て
            match_damage_min_cr = math.floor(sametype_damage_min_cr * 0.5)
        elif ws["B5"].value == "かなりいまひとつ":
            # 最大
            # 切り捨て
            match_damage_max = math.floor(sametype_damage_max * 0.25)
            # 最小
            # 切り捨て
            match_damage_min = math.floor(sametype_damage_min * 0.25)

            # 急所
            # 最大
            # 切り捨て
            match_damage_max_cr = math.floor(sametype_damage_max_cr * 0.25)
            # 最小
            # 切り捨て
            match_damage_min_cr = math.floor(sametype_damage_min_cr * 0.25)
        elif ws["B5"].value == "こうかなし":
            # 最大
            match_damage_max = sametype_damage_max * 0
            # 最小
            match_damage_min = sametype_damage_min * 0

            # 急所
            # 最大
            # 切り捨て
            match_damage_max_cr = sametype_damage_max_cr * 0
            # 最小
            # 切り捨て
            match_damage_min_cr = sametype_damage_min_cr * 0
        else: # 等倍
            # 最大
            match_damage_max = sametype_damage_max
            # 最小
            match_damage_min = sametype_damage_min

            # 急所
            # 最大
            # 切り捨て
            match_damage_max_cr = sametype_damage_max_cr
            # 最小
            # 切り捨て
            match_damage_min_cr = sametype_damage_min_cr
        
        # 残り体力計算
        rem_HP_max = HP_act - match_damage_max
        rem_HP_min = HP_act - match_damage_min

        # 急所
        rem_HP_max_cr = HP_act - match_damage_max_cr
        rem_HP_min_cr = HP_act - match_damage_min_cr


        self.loading_message["text"] = "ダメージ", match_damage_max, match_damage_min, "残り体力", rem_HP_max, rem_HP_min, "急所", match_damage_max_cr, match_damage_min_cr, "残り体力", rem_HP_max_cr, rem_HP_min_cr


    # リセットボタン機能
    def reset(self):
        for i in self.cell_texts:
            i.delete(0, tkinter.END)
        self.checked_type_same.set(0)
        self.text_box_type_match.set(0)
        
        wb = openpyxl.load_workbook("app_data.xlsx") # 選択したファイルを読み込み
        ws = wb.worksheets[0] # ワークシートの一枚目(0番目)をwsに保存
        text_A_C = self.text_box_A_C.get()
        ws["A1"].value = text_A_C
        text_efA_C = self.text_box_efA_C.get()
        ws["A2"].value = text_efA_C
        text_mo = self.text_box_mo.get()
        ws["A3"].value = text_mo
        check_box_type_same = self.checked_type_same.get()
        ws["A4"].value = check_box_type_same


        text_B_D = self.text_box_B_D.get()
        ws["B1"].value = text_B_D
        text_HP = self.text_box_HP.get()
        ws["B2"].value = text_HP
        text_efB_D = self.text_box_efB_D.get()
        ws["B3"].value = text_efB_D
        text_efHP = self.text_box_efHP.get()
        ws["B4"].value = text_efHP
        text_box_type_match = self.text_box_type_match.get()
        ws["B5"].value = text_box_type_match
        
        
        wb.save("app_data.xlsx") # book保存
        self.message["text"] = "リセット"
    
    # 読み込みボタンを押すとエクセル内容が表示されるメソッドを作成
    def loading_data(self):
        wb = openpyxl.load_workbook("app_data.xlsx")
        ws = wb.worksheets[0] 
        values = [] # cellの中の文字を入れるためのリスト
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None: # cellの中に文字がある時にリストにその文字を追加する
                    values.append(cell.value)
        self.loading_message["text"] = values

root = tkinter.Tk()
root.title("ダメージ計算 アプリ")
root.geometry("850x1150")
app = Application(root=root)
root.mainloop()